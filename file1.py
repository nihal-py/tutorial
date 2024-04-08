import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import openpyxl

def send_email(subject, body):
    try:
        email = 'retrohubmusic@gmail.com'
        password = 'howktwkhtnmmokgq'
        send_to_email = 'nihal.techworks@gmail.com'

        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = send_to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email, password)
        text = msg.as_string()
        server.sendmail(email, send_to_email, text)
        server.quit()
        print("Email sent successfully")

        # Save email details to Excel
        save_email_to_excel(email, send_to_email, subject, body)

    except Exception as e:
        print(f"Error sending email: {e}")

def save_email_to_excel(sender, receiver, subject, content):
    try:
        wb = openpyxl.load_workbook('sent_emails.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active

    # Add headers if not already present
    if not ws['A1'].value:
        ws['A1'] = 'From'
        ws['B1'] = 'To'
        ws['C1'] = 'Date'
        ws['D1'] = 'Time'
        ws['E1'] = 'Subject'
        ws['F1'] = 'Content'

    # Append new data
    row = (sender, receiver, datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%H:%M:%S'), subject, content)
    ws.append(row)

    # Save to file
    wb.save('sent_emails.xlsx')

# Test the function
send_email("Test Subject", "This is a test email content.")
