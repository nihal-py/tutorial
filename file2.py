import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import openpyxl
import tkinter as tk
from tkinter import messagebox

def send_email():
    try:
        email = 'retrohubmusic@gmail.com'
        password = 'howktwkhtnmmokgq'
        send_to_email = receiver_entry.get()
        subject = subject_entry.get()
        body = content_entry.get("1.0", "end-1c")

        msg = MIMEMultipart()
        msg['From'] = email
        msg['To'] = send_to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(email, password)
        text = msg.as_string()
        server.sendmail(email, send_to_email, text)
        server.quit()
        print("Email sent successfully")

        # Save email details to Excel
        save_email_to_excel(email, send_to_email, subject, body)
        
        messagebox.showinfo("Success", "Email sent successfully")

    except Exception as e:
        print(f"Error sending email: {e}")
        messagebox.showerror("Error", f"Error sending email: {e}")

def save_email_to_excel(sender, receiver, subject, content):
    try:
        wb = openpyxl.load_workbook('sent_emails.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    ws = wb.active

    # Add headers if not already present
    if not ws['A1'].value:
        ws['A1'] = 'From'
        ws['B1'] = 'To'
        ws['C1'] = 'Date'
        ws['D1'] = 'Time'
        ws['E1'] = 'Subject'
        ws['F1'] = 'Content'

    # Append new data
    row = (sender, receiver, datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%H:%M:%S'), subject, content)
    ws.append(row)

    # Save to file
    wb.save('sent_emails.xlsx')

# Create GUI
root = tk.Tk()
root.title("Email Sender")

receiver_label = tk.Label(root, text="Receiver's Email:")
receiver_label.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)

receiver_entry = tk.Entry(root, width=40)
receiver_entry.grid(row=0, column=1, padx=10, pady=5)

subject_label = tk.Label(root, text="Subject:")
subject_label.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)

subject_entry = tk.Entry(root, width=40)
subject_entry.grid(row=1, column=1, padx=10, pady=5)

content_label = tk.Label(root, text="Content:")
content_label.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)

content_entry = tk.Text(root, width=30, height=10)
content_entry.grid(row=2, column=1, padx=10, pady=5)

send_button = tk.Button(root, text="Send", command=send_email)
send_button.grid(row=3, column=1, padx=10, pady=5, sticky=tk.E)

root.mainloop()
